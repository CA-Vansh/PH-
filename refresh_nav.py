#!/usr/bin/env python3
"""
refresh_nav.py — MF NAV & Returns Master Workbook

Usage
-----
  python refresh_nav.py --healthcheck            Test both data sources, exit
  python refresh_nav.py --init                   Create an empty scaffold workbook
  python refresh_nav.py --resolve --names-file f Look up scheme codes for a list of funds
  python refresh_nav.py                          Refresh using the scheme codes tab
  python refresh_nav.py --amfi-history           Allow AMFI history fallback (slow)
  python refresh_nav.py --selftest               Offline logic tests (no network)

Data sources
------------
  Primary  : mfapi.in   https://api.mfapi.in/mf/{code}          full NAV history
  Fallback1: AMFI snap  https://portal.amfiindia.com/spages/NAVAll.txt
             Current-day snapshot only. No history.
  Fallback2: AMFI hist  DownloadNAVHistoryReport_Po.aspx?frmdt=&todt=
             Real history, but capped at 90 DAYS PER REQUEST. Off by default
             (--amfi-history) because rebuilding a 3Y anchor needs ~13 calls
             of a multi-MB report each.

  IMPORTANT (verified 24-Aug-2026): AMFI migrated to portal.amfiindia.com and
  changed the text format. Old 6-column layout retires 28-Aug-2026. This parser
  auto-detects both:
      OLD 6 cols: Code;ISIN1;ISIN2;Name;NAV;Date
      NEW 8 cols: Code;ISIN1;ISIN2;Name;Plan;Option;NAV;Date

Tabs
----
  Scheme Codes  INPUT ONLY  — paste codes here; nothing else to fill in
  NAV Anchors   SCRIPT-WRITTEN — fund name/AMC/category auto-filled from source
  NAVs          SCRIPT-WRITTEN — anchor points + latest, long format
  Returns       LIVE FORMULAS ONLY, referencing 'NAV Anchors'
  Refresh Log   SCRIPT-WRITTEN append-only run history
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from statistics import median

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_FILE = "MF_NAV_Master.xlsx"
MFAPI_URL = "https://api.mfapi.in/mf/{code}"
AMFI_SNAPSHOT_URL = "https://portal.amfiindia.com/spages/NAVAll.txt"
AMFI_HISTORY_URL = ("https://portal.amfiindia.com/DownloadNAVHistoryReport_Po.aspx"
                    "?frmdt={frm}&todt={to}")
AMFI_HISTORY_MAX_DAYS = 90        # hard limit imposed by AMFI

REQUEST_DELAY = 0.25
MAX_RETRIES = 4
BACKOFF_BASE = 1.6
TIMEOUT = 30

ANCHORS = [("1M", 1), ("3M", 3), ("6M", 6), ("1Y", 12), ("3Y", 36), ("Earliest", None)]

SPIKE_WINDOW = 5
SPIKE_TOLERANCE = 0.25

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
BLUE = Font(name=FONT, color="0000FF", size=10)
BLACK = Font(name=FONT, color="000000", size=10)
GREEN = Font(name=FONT, color="008000", size=10)
BASE = Font(name=FONT, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FMT = "DD-MMM-YYYY"
NAV_FMT = "#,##0.0000"
PCT_FMT = "0.00%"

# ---- layouts ---------------------------------------------------------------

WATCHLIST_COLS = [("Scheme Code", 14), ("Your Label (optional)", 40), ("Include? (Y/N)", 14)]
WL_HEADER_ROW = 4
WL_FIRST_DATA_ROW = 5

# A code | B name | C AMC | D category | E latest date | F latest NAV
# then anchor (date,nav) pairs from col 7, then source + flag
ANCHOR_COLS = ["Scheme Code", "Fund Name", "AMC", "Category", "Latest Date", "Latest NAV"]
for _lbl, _ in ANCHORS:
    ANCHOR_COLS += [f"{_lbl} Date", f"{_lbl} NAV"]
ANCHOR_COLS += ["Data Source", "Quality Flag"]
AN_HEADER_ROW = 3
AN_FIRST_DATA_ROW = 4

AN_NAME = get_column_letter(2)
AN_AMC = get_column_letter(3)
AN_CAT = get_column_letter(4)
AN_LATEST_DATE = get_column_letter(5)
AN_LATEST_NAV = get_column_letter(6)
AN_PAIR = {lbl: (get_column_letter(7 + i * 2), get_column_letter(8 + i * 2))
           for i, (lbl, _) in enumerate(ANCHORS)}
AN_SOURCE_COL = 7 + len(ANCHORS) * 2
AN_FLAG_COL = 8 + len(ANCHORS) * 2

RETURNS_COLS = [("Scheme Code", 13), ("Fund Name", 46), ("AMC", 22), ("Category", 26),
                ("1M", 10), ("3M", 10), ("6M", 10), ("1Y", 10), ("3Y CAGR", 12),
                ("Since Earliest (CAGR)", 20), ("Earliest NAV Date", 18),
                ("Years Covered", 13)]
RT_HEADER_ROW = 3
RT_FIRST_DATA_ROW = 4

LOG_COLS = [("Run Timestamp", 20), ("Scheme Code", 13), ("Fund Name", 42),
            ("Status", 12), ("Data Source", 22), ("Records", 10),
            ("Latest NAV Date", 17), ("Message", 60)]
LOG_HEADER_ROW = 3
LOG_FIRST_DATA_ROW = 4


# ---- pure logic (covered by --selftest) ------------------------------------

def months_back(d: date, months: int) -> date:
    y, m = d.year, d.month - months
    while m <= 0:
        m += 12
        y -= 1
    day = d.day
    while day > 0:
        try:
            return date(y, m, day)
        except ValueError:
            day -= 1
    raise ValueError("unreachable")


def parse_history(payload: dict) -> list[tuple[date, float]]:
    out = []
    for row in payload.get("data", []):
        try:
            nav = float((row.get("nav") or "").strip())
        except (TypeError, ValueError):
            continue
        if nav <= 0:
            continue
        try:
            d = datetime.strptime(row["date"].strip(), "%d-%m-%Y").date()
        except (KeyError, ValueError, AttributeError):
            continue
        out.append((d, nav))
    out.sort(key=lambda x: x[0])
    return out


def parse_amfi_line(line: str):
    """Handle BOTH AMFI layouts. Returns dict or None.

    OLD 6 cols: Code;ISIN1;ISIN2;Name;NAV;Date
    NEW 8 cols: Code;ISIN1;ISIN2;Name;Plan;Option;NAV;Date
    NAV and Date are always the LAST TWO fields, so index from the end.
    """
    parts = [p.strip() for p in line.split(";")]
    if len(parts) < 6 or not parts[0].isdigit():
        return None
    try:
        nav = float(parts[-2])
        d = datetime.strptime(parts[-1], "%d-%b-%Y").date()
    except ValueError:
        return None
    if nav <= 0:
        return None
    plan = parts[4] if len(parts) >= 8 else ""
    option = parts[5] if len(parts) >= 8 else ""
    return {"code": parts[0], "name": parts[3], "plan": plan,
            "option": option, "nav": nav, "date": d}


def pick_anchor(series, target: date):
    chosen = None
    for d, nav in series:
        if d <= target:
            chosen = (d, nav)
        else:
            break
    return chosen


def spike_flag(series, anchor) -> bool:
    if not anchor:
        return False
    idx = next((i for i, (d, _) in enumerate(series) if d == anchor[0]), None)
    if idx is None:
        return False
    lo, hi = max(0, idx - SPIKE_WINDOW), min(len(series), idx + SPIKE_WINDOW + 1)
    neigh = [n for i, (_, n) in enumerate(series[lo:hi], start=lo) if i != idx]
    if len(neigh) < 3:
        return False
    med = median(neigh)
    return med > 0 and abs(anchor[1] - med) / med > SPIKE_TOLERANCE


def build_anchor_set(series) -> dict:
    if not series:
        return {}
    latest_d, latest_n = series[-1]
    res = {"latest": (latest_d, latest_n), "flags": []}
    for label, months in ANCHORS:
        hit = series[0] if months is None else pick_anchor(series, months_back(latest_d, months))
        res[label] = hit
        if hit and spike_flag(series, hit):
            res["flags"].append(f"{label} NAV looks like an outlier")
    return res


def is_idcw(text: str) -> bool:
    s = (text or "").upper()
    return any(t in s for t in ("IDCW", "DIVIDEND", "PAYOUT", "REINVEST", "RE-INVEST", "DCW"))


MFAPI_SEARCH_URL = "https://api.mfapi.in/mf/search?q={q}"

# Common shorthand seen in client sheets / factsheets, expanded for matching.
_ABBREV = [
    (r"\bdirect\b", "direct"), (r"\bdir\b", "direct"),
    (r"\breg(ular)?\b", "regular"),
    (r"\bg\b", "growth"), (r"\bgr\b", "growth"), (r"\bgrowth\b", "growth"),
    (r"\bidcw\b", "idcw"), (r"\bdiv(idend)?\b", "idcw"),
    (r"\bsmallcap\b", "small cap"), (r"\bmidcap\b", "mid cap"),
    (r"\blargecap\b", "large cap"), (r"\bflexicap\b", "flexi cap"),
    (r"\bmulticap\b", "multi cap"),
    (r"[-_/]", " "),
    # NOTE: do NOT expand "teck" -> "tech": Quant Teck Fund is the fund's
    # actual, correct name (not a typo), so rewriting it breaks the search.
]


def normalize_name(name: str) -> str:
    """Lowercase, expand shorthand, collapse whitespace — for fuzzy matching only."""
    s = name.lower()
    for pat, repl in _ABBREV:
        s = re.sub(pat, repl, s)
    return re.sub(r"\s+", " ", s).strip()


def search_query_terms(name: str) -> str:
    """Strip plan/option words to build a broad mfapi search query.

    mfapi's search matches on the scheme name string; plan/option words like
    'direct' or 'growth' are common to thousands of schemes and dilute the
    match, so the query itself should just be the fund's identifying words.
    """
    norm = normalize_name(name)
    drop = {"direct", "regular", "growth", "idcw", "plan", "fund", "option"}
    kept = [w for w in norm.split() if w not in drop]
    return " ".join(kept) if kept else norm


# mfapi's search does literal substring matching against the AMC's actual
# scheme name, and AMCs are NOT consistent about spacing: HSBC files "Mid
# Cap" (two words), Invesco and Motilal Oswal file "Smallcap"/"Midcap" (one
# word). normalize_name always expands to the spaced form for comparison,
# but the SEARCH QUERY itself needs both spellings tried, or a fund whose
# AMC uses the one-word convention returns zero results even though it's
# sitting right there in the index.
_COLLAPSE = [("small cap", "smallcap"), ("mid cap", "midcap"),
            ("large cap", "largecap"), ("flexi cap", "flexicap"),
            ("multi cap", "multicap")]


def search_query_variants(name: str) -> list[str]:
    """One or two query strings to try: spaced form and one-word form."""
    spaced = search_query_terms(name)
    collapsed = spaced
    for two_word, one_word in _COLLAPSE:
        collapsed = collapsed.replace(two_word, one_word)
    return [spaced] if collapsed == spaced else [spaced, collapsed]


def score_candidate(target_name: str, candidate_name: str) -> float:
    """Similarity between the requested fund (with plan/option) and a result."""
    return SequenceMatcher(None, normalize_name(target_name),
                           normalize_name(candidate_name)).ratio()


def rank_candidates(target_name: str, candidates: list[dict], top_n: int = 5) -> list[dict]:
    """candidates: [{'schemeCode':.., 'schemeName':..}, ...] from mfapi /mf/search.
    Returns them sorted by similarity to target_name, annotated with 'score'.
    Never picks a winner — ranking is for a human to read and confirm.
    """
    out = []
    for c in candidates:
        code = c.get("schemeCode") or c.get("scheme_code")
        cname = c.get("schemeName") or c.get("scheme_name") or ""
        if code is None:
            continue
        out.append({"code": str(code), "name": cname,
                    "score": score_candidate(target_name, cname)})
    out.sort(key=lambda x: x["score"], reverse=True)
    return out[:top_n]


def chunk_ranges(start: date, end: date, max_days: int = AMFI_HISTORY_MAX_DAYS):
    """Split [start,end] into <=max_days windows (AMFI's hard cap)."""
    out, cur = [], start
    while cur <= end:
        stop = min(cur + timedelta(days=max_days - 1), end)
        out.append((cur, stop))
        cur = stop + timedelta(days=1)
    return out


# ---- network ---------------------------------------------------------------

def _requests():
    try:
        import requests
        return requests
    except ImportError:
        sys.exit("requests not installed. Run: pip install requests openpyxl")


def fetch_mfapi(code: str) -> dict:
    requests = _requests()
    last = None
    for attempt in range(MAX_RETRIES):
        try:
            r = requests.get(MFAPI_URL.format(code=code), timeout=TIMEOUT)
            if r.status_code == 429 or r.status_code >= 500:
                raise RuntimeError(f"HTTP {r.status_code}")
            r.raise_for_status()
            payload = r.json()
            if not payload.get("data"):
                raise RuntimeError("empty data array")
            return payload
        except Exception as exc:                     # noqa: BLE001
            last = exc
            if attempt < MAX_RETRIES - 1:
                time.sleep(BACKOFF_BASE ** attempt)
    raise RuntimeError(f"mfapi failed after {MAX_RETRIES} tries: {last}")


_snapshot_cache = None


def fetch_amfi_snapshot() -> dict:
    global _snapshot_cache
    if _snapshot_cache is not None:
        return _snapshot_cache
    requests = _requests()
    r = requests.get(AMFI_SNAPSHOT_URL, timeout=TIMEOUT)
    r.raise_for_status()
    table = {}
    for line in r.text.splitlines():
        rec = parse_amfi_line(line)
        if rec:
            table[rec["code"]] = rec
    if not table:
        raise RuntimeError("NAVAll.txt parsed to zero rows — format may have changed again")
    _snapshot_cache = table
    return table


def fetch_amfi_history(code: str, start: date, end: date) -> list[tuple[date, float]]:
    """Rebuild a scheme's history from AMFI in <=90-day chunks."""
    requests = _requests()
    series = []
    windows = chunk_ranges(start, end)
    for i, (frm, to) in enumerate(windows, start=1):
        url = AMFI_HISTORY_URL.format(frm=frm.strftime("%d-%b-%Y"), to=to.strftime("%d-%b-%Y"))
        print(f"      AMFI history window {i}/{len(windows)}: {frm} -> {to}")
        r = requests.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        for line in r.text.splitlines():
            rec = parse_amfi_line(line)
            if rec and rec["code"] == code:
                series.append((rec["date"], rec["nav"]))
        time.sleep(REQUEST_DELAY)
    series = sorted(set(series), key=lambda x: x[0])
    return series


def fetch_mfapi_search(query: str) -> list[dict]:
    requests = _requests()
    r = requests.get(MFAPI_SEARCH_URL.format(q=query), timeout=TIMEOUT)
    r.raise_for_status()
    data = r.json()
    return data if isinstance(data, list) else []


def resolve_names(names: list[str], top_n: int = 5) -> list[dict]:
    """For each requested fund name, search mfapi.in and rank the results.
    Tries both spaced and one-word forms of cap-related terms and merges
    results, since AMCs spell "Smallcap"/"Small Cap" inconsistently.
    Never auto-selects — every row needs human confirmation.
    """
    report = []
    for raw in names:
        raw = raw.strip()
        if not raw:
            continue
        variants = search_query_variants(raw)
        merged, seen, status = [], set(), "ok"
        try:
            for q in variants:
                for c in fetch_mfapi_search(q):
                    code = c.get("schemeCode") or c.get("scheme_code")
                    if code is not None and code not in seen:
                        seen.add(code)
                        merged.append(c)
                time.sleep(REQUEST_DELAY)
            ranked = rank_candidates(raw, merged, top_n=top_n)
            if not ranked:
                status = "no_results"
        except Exception as exc:                      # noqa: BLE001
            ranked, status = [], f"error: {exc}"
        report.append({"requested": raw, "query": " / ".join(variants),
                       "status": status, "candidates": ranked})
    return report


def print_resolve_report(report: list[dict]):
    print("\nScheme code resolution — REVIEW BEFORE PASTING ANYWHERE")
    print("=" * 78)
    for row in report:
        print(f"\n  Requested : {row['requested']}")
        print(f"  Searched  : \"{row['query']}\"")
        if row["status"] == "no_results":
            print("  NO MATCHES — check spelling, or the fund may not be in mfapi's index")
            continue
        if row["status"] != "ok":
            print(f"  ERROR: {row['status']}")
            continue
        best = row["candidates"][0]["score"]
        for i, c in enumerate(row["candidates"]):
            flag = " <-- best match" if i == 0 else ""
            weak = "  (LOW CONFIDENCE)" if c["score"] < 0.55 else ""
            print(f"    [{c['score']:.0%}] {c['code']:>8}  {c['name']}{flag}{weak}")
        if best < 0.55:
            print("  ! No strong match — verify manually on mfapi.in or AMFI before using.")
    print("\n" + "=" * 78)
    print("Nothing has been written to the workbook. Copy the CORRECT code for each")
    print("fund into the 'Scheme Codes' tab yourself, then run: python refresh_nav.py")


def healthcheck() -> int:
    """Probe both sources and report exactly what works."""
    requests = _requests()
    problems = 0
    print("Source health check\n" + "=" * 60)

    print("\n[1/3] mfapi.in — history endpoint")
    try:
        t0 = time.time()
        r = requests.get(MFAPI_URL.format(code=119551), timeout=TIMEOUT)
        ms = (time.time() - t0) * 1000
        print(f"      HTTP {r.status_code} in {ms:.0f} ms")
        payload = r.json()
        series = parse_history(payload)
        meta = payload.get("meta", {})
        print(f"      scheme  : {meta.get('scheme_name', '?')[:60]}")
        print(f"      records : {len(series)}")
        if series:
            print(f"      latest  : {series[-1][0]} = {series[-1][1]}")
            age = (date.today() - series[-1][0]).days
            print(f"      staleness: {age} day(s)")
            if age > 5:
                print("      WARNING: latest NAV is unusually old")
        print("      OK" if series else "      FAIL: no usable rows")
        problems += 0 if series else 1
    except Exception as exc:                          # noqa: BLE001
        print(f"      FAIL: {exc}")
        problems += 1

    print("\n[2/3] AMFI snapshot — NAVAll.txt")
    try:
        t0 = time.time()
        r = requests.get(AMFI_SNAPSHOT_URL, timeout=TIMEOUT)
        ms = (time.time() - t0) * 1000
        header = r.text.splitlines()[0] if r.text else ""
        ncols = len(header.split(";"))
        print(f"      HTTP {r.status_code} in {ms:.0f} ms, {len(r.text) // 1024} KB")
        print(f"      header cols: {ncols} "
              f"({'NEW 8-col' if ncols >= 8 else 'OLD 6-col'} layout)")
        table = {}
        for line in r.text.splitlines():
            rec = parse_amfi_line(line)
            if rec:
                table[rec["code"]] = rec
        print(f"      parsed schemes: {len(table)}")
        if table:
            sample = next(iter(table.values()))
            print(f"      sample: {sample['code']} {sample['nav']} on {sample['date']}")
        print("      OK" if table else "      FAIL: parsed zero rows")
        problems += 0 if table else 1
    except Exception as exc:                          # noqa: BLE001
        print(f"      FAIL: {exc}")
        problems += 1

    print("\n[3/3] AMFI history — DownloadNAVHistoryReport_Po.aspx")
    try:
        end = date.today()
        start = end - timedelta(days=10)
        url = AMFI_HISTORY_URL.format(frm=start.strftime("%d-%b-%Y"),
                                      to=end.strftime("%d-%b-%Y"))
        t0 = time.time()
        r = requests.get(url, timeout=TIMEOUT)
        ms = (time.time() - t0) * 1000
        rows = sum(1 for ln in r.text.splitlines() if parse_amfi_line(ln))
        print(f"      HTTP {r.status_code} in {ms:.0f} ms, {len(r.text) // 1024} KB")
        print(f"      parsed rows (10-day window): {rows}")
        print("      OK" if rows else "      FAIL: parsed zero rows")
        problems += 0 if rows else 1
    except Exception as exc:                          # noqa: BLE001
        print(f"      FAIL: {exc}")
        problems += 1

    print("\n" + "=" * 60)
    print("All sources reachable." if not problems else f"{problems} source(s) FAILED.")
    print("Reminder: AMFI's OLD 6-column format retires 28-Aug-2026.")
    return 1 if problems else 0


# ---- styling ---------------------------------------------------------------

def style_header(ws, row, cols):
    for i, item in enumerate(cols, start=1):
        title, width = item if isinstance(item, tuple) else (item, 16)
        c = ws.cell(row=row, column=i, value=title)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def title_block(ws, title, subtitle):
    ws.cell(row=1, column=1, value=title).font = Font(name=FONT, bold=True, size=13,
                                                      color="1F3864")
    ws.cell(row=2, column=1, value=subtitle).font = Font(name=FONT, italic=True, size=9,
                                                         color="595959")


# ---- scaffold --------------------------------------------------------------

def build_scaffold(path: str):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Scheme Codes"
    title_block(ws, "Scheme Codes — INPUT ONLY",
                "Paste mfapi.in / AMFI scheme codes. That is all you fill in — fund name, "
                "AMC and category are pulled automatically from the data source.")
    ws.cell(row=3, column=1,
            value="Tip: a scheme code identifies ONE plan and ONE option. Growth and IDCW "
                  "have different codes, so pick the code for the exact variant you track."
            ).font = Font(name=FONT, italic=True, size=9, color="C00000")
    style_header(ws, WL_HEADER_ROW, WATCHLIST_COLS)

    for i, val in enumerate(["119550", "ABSL Banking & PSU — Direct Growth", "Y"], start=1):
        c = ws.cell(row=WL_FIRST_DATA_ROW, column=i, value=val)
        c.font, c.fill, c.border = BLUE, INPUT_FILL, BORDER
    ws.cell(row=WL_FIRST_DATA_ROW, column=1).comment = Comment(
        "EXAMPLE ROW — overwrite or delete.\n"
        "119550 is the Direct-Growth variant. 119551 is the same fund's "
        "IDCW-Reinvestment variant, whose NAV drops on every payout and would "
        "make price-only returns look negative.", "refresh_nav.py")

    for r in range(WL_FIRST_DATA_ROW + 1, WL_FIRST_DATA_ROW + 100):
        for i in range(1, len(WATCHLIST_COLS) + 1):
            c = ws.cell(row=r, column=i)
            c.fill, c.font, c.border = INPUT_FILL, BLUE, BORDER

    lr = WL_FIRST_DATA_ROW + 102
    ws.cell(row=lr, column=1, value="LEGEND").font = Font(name=FONT, bold=True, size=10)
    for off, txt in enumerate([
        "Yellow + blue text = you type here. This is the only tab you edit.",
        "Black text = calculated by a formula.  Green text = link to another tab.",
        "'NAV Anchors', 'NAVs', 'Refresh Log' are rebuilt each run — do not hand-edit.",
        "'Returns' is live formulas only. Editing a cell there breaks self-updating.",
        "Blank the Include? cell or set N to park a code without deleting the row.",
    ], start=1):
        ws.cell(row=lr + off, column=1, value=txt).font = Font(name=FONT, size=9)

    wa = wb.create_sheet("NAV Anchors")
    title_block(wa, "NAV Anchors — SCRIPT-WRITTEN VALUES",
                "Reference NAVs at fixed lookbacks. Anchor dates carry BACK to the last "
                "traded day, never forward. Rebuilt every run.")
    style_header(wa, AN_HEADER_ROW,
                 [(c, 40 if c in ("Fund Name", "Quality Flag") else 16) for c in ANCHOR_COLS])

    wn = wb.create_sheet("NAVs")
    title_block(wn, "NAVs — SCRIPT-WRITTEN VALUES",
                "Anchor points + latest NAV per scheme, long format. Rebuilt every run.")
    style_header(wn, 3, [("Scheme Code", 13), ("Fund Name", 46), ("Point", 12),
                         ("NAV Date", 15), ("NAV", 14), ("Source", 22)])

    wr = wb.create_sheet("Returns")
    title_block(wr, "Returns — LIVE FORMULAS ONLY",
                "Every cell references 'NAV Anchors'. Simple returns to 1Y; CAGR for 3Y and "
                "since-earliest, annualised on ACTUAL elapsed days. All wrapped in IFERROR.")
    style_header(wr, RT_HEADER_ROW, RETURNS_COLS)

    wl = wb.create_sheet("Refresh Log")
    title_block(wl, "Refresh Log — SCRIPT-WRITTEN, APPEND-ONLY",
                "One row per scheme per run. 'amfi-snapshot' = latest NAV only, anchors "
                "not rebuilt. 'amfi-history' = anchors rebuilt from 90-day chunks.")
    style_header(wl, LOG_HEADER_ROW, LOG_COLS)

    wb.save(path)
    return path


# ---- refresh ---------------------------------------------------------------

def read_codes(ws):
    out, r, blanks = [], WL_FIRST_DATA_ROW, 0
    while blanks < 30:
        code = ws.cell(row=r, column=1).value
        if not code:
            blanks += 1
            r += 1
            continue
        blanks = 0
        inc = str(ws.cell(row=r, column=3).value or "Y").strip().upper()
        if not inc.startswith("N"):
            out.append({"code": str(code).strip().split(".")[0],
                        "label": ws.cell(row=r, column=2).value or "", "row": r})
        r += 1
    return out


def clear_below(ws, first_row):
    if ws.max_row >= first_row:
        ws.delete_rows(first_row, ws.max_row - first_row + 2)


def write_returns_formulas(wr, n_rows):
    SRC = "'NAV Anchors'"
    for i in range(n_rows):
        row, arow = RT_FIRST_DATA_ROW + i, AN_FIRST_DATA_ROW + i
        for col, ref in ((1, "A"), (2, AN_NAME), (3, AN_AMC), (4, AN_CAT)):
            c = wr.cell(row=row, column=col,
                        value=f'=IFERROR(IF({SRC}!{ref}{arow}="","-",{SRC}!{ref}{arow}),"-")')
            c.font, c.border = GREEN, BORDER

        lat_d, lat_n = f"{SRC}!{AN_LATEST_DATE}{arow}", f"{SRC}!{AN_LATEST_NAV}{arow}"

        for col, label in ((5, "1M"), (6, "3M"), (7, "6M"), (8, "1Y")):
            anch = f"{SRC}!{AN_PAIR[label][1]}{arow}"
            c = wr.cell(row=row, column=col,
                        value=f'=IFERROR(IF(OR({anch}="",{lat_n}=""),"-",{lat_n}/{anch}-1),"-")')
            c.number_format, c.font, c.border = PCT_FMT, BLACK, BORDER

        for col, label in ((9, "3Y"), (10, "Earliest")):
            dcol, ncol = AN_PAIR[label]
            an, ad = f"{SRC}!{ncol}{arow}", f"{SRC}!{dcol}{arow}"
            c = wr.cell(row=row, column=col,
                        value=f'=IFERROR(IF(OR({an}="",{ad}="",{lat_n}="",{lat_d}-{ad}<=0),'
                              f'"-",({lat_n}/{an})^(365.25/({lat_d}-{ad}))-1),"-")')
            c.number_format, c.font, c.border = PCT_FMT, BLACK, BORDER

        ed = f"{SRC}!{AN_PAIR['Earliest'][0]}{arow}"
        c = wr.cell(row=row, column=11, value=f'=IFERROR(IF({ed}="","-",{ed}),"-")')
        c.number_format, c.font, c.border = DATE_FMT, GREEN, BORDER
        c = wr.cell(row=row, column=12,
                    value=f'=IFERROR(IF(OR({ed}="",{lat_d}=""),"-",({lat_d}-{ed})/365.25),"-")')
        c.number_format, c.font, c.border = "0.00", BLACK, BORDER


def refresh_workbook(wb, allow_amfi_history: bool = False, on_progress=None):
    """Core refresh logic, operating on an already-open Workbook object.

    Does NOT load or save any file — caller owns that (path on disk, or
    BytesIO for the Streamlit app). Returns (wb, ok, fail).

    on_progress, if given, is called as on_progress(i, total, code, status, name)
    after each scheme code is processed, instead of printing to stdout —
    used by the Streamlit UI to drive a live progress bar/table.
    """
    for tab in ("Scheme Codes", "NAV Anchors", "NAVs", "Returns", "Refresh Log"):
        if tab not in wb.sheetnames:
            raise ValueError(f"Missing tab '{tab}'. Use --init / a fresh scaffold.")

    codes = read_codes(wb["Scheme Codes"])
    if not codes:
        raise ValueError("No scheme codes found. Add codes to the 'Scheme Codes' tab.")

    wa, wn, wr, wl = wb["NAV Anchors"], wb["NAVs"], wb["Returns"], wb["Refresh Log"]
    for ws, first in ((wa, AN_FIRST_DATA_ROW), (wn, 4), (wr, RT_FIRST_DATA_ROW)):
        clear_below(ws, first)

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_row, nav_row = max(wl.max_row + 1, LOG_FIRST_DATA_ROW), 4
    ok = fail = 0

    for i, f in enumerate(codes):
        arow = AN_FIRST_DATA_ROW + i
        code = f["code"]
        source = status = message = ""
        name, amc, category = f["label"] or code, "", ""
        anchors, records, latest, flags = {}, 0, None, []

        try:
            payload = fetch_mfapi(code)
            series = parse_history(payload)
            if not series:
                raise RuntimeError("mfapi returned no usable NAV rows")
            meta = payload.get("meta", {})
            name = meta.get("scheme_name") or name
            amc = meta.get("fund_house", "")
            category = meta.get("scheme_category", "")
            anchors = build_anchor_set(series)
            records, latest = len(series), anchors["latest"]
            source, status = "mfapi.in", "success"
            flags = list(anchors.get("flags", []))
            time.sleep(REQUEST_DELAY)

        except Exception as exc:                      # noqa: BLE001
            try:
                snap = fetch_amfi_snapshot()
                rec = snap.get(code)
                if not rec:
                    raise RuntimeError(f"not in NAVAll.txt either ({exc})")
                name = rec["name"] or name
                if rec["plan"] or rec["option"]:
                    name = f"{rec['name']} - {rec['plan']} - {rec['option']}".strip(" -")
                latest = (rec["date"], rec["nav"])

                if allow_amfi_history:
                    hist = fetch_amfi_history(code, latest[0] - timedelta(days=365 * 3 + 5),
                                              latest[0])
                    if hist:
                        anchors = build_anchor_set(hist)
                        records, latest = len(hist), anchors["latest"]
                        source, status = "amfi-history", "success"
                        flags = list(anchors.get("flags", []))
                        flags.append("anchors rebuilt from AMFI; 'Earliest' is capped at "
                                     "the 3Y window, NOT true inception")
                    else:
                        raise RuntimeError("AMFI history returned nothing for this code")
                else:
                    anchors, records = {"latest": latest}, 1
                    source, status = "amfi-snapshot", "partial"
                    flags = [f"mfapi failed ({str(exc)[:80]}). Snapshot gave latest NAV "
                             f"only; anchors NOT rebuilt. Re-run with --amfi-history."]
            except Exception as exc2:                 # noqa: BLE001
                source, status = "none", "fail"
                message = str(exc2)[:250]
                fail += 1

        if status in ("success", "partial"):
            ok += 1
            if is_idcw(name):
                flags.append("IDCW/payout variant — price-only returns exclude distributions")
            if latest and (date.today() - latest[0]).days > 7:
                flags.append(f"latest NAV is {(date.today() - latest[0]).days} days old")
            message = "; ".join(flags)

        wa.cell(row=arow, column=1, value=code).font = BASE
        wa.cell(row=arow, column=2, value=name).font = BASE
        wa.cell(row=arow, column=3, value=amc).font = BASE
        wa.cell(row=arow, column=4, value=category).font = BASE
        if latest:
            c = wa.cell(row=arow, column=5, value=latest[0]); c.number_format = DATE_FMT; c.font = BASE
            c = wa.cell(row=arow, column=6, value=latest[1]); c.number_format = NAV_FMT; c.font = BASE
        for j, (label, _) in enumerate(ANCHORS):
            hit = anchors.get(label)
            if hit:
                c = wa.cell(row=arow, column=7 + j * 2, value=hit[0])
                c.number_format, c.font = DATE_FMT, BASE
                c = wa.cell(row=arow, column=8 + j * 2, value=hit[1])
                c.number_format, c.font = NAV_FMT, BASE
        wa.cell(row=arow, column=AN_SOURCE_COL, value=source).font = BASE
        fc = wa.cell(row=arow, column=AN_FLAG_COL, value=message)
        fc.font = Font(name=FONT, size=10, color="C00000" if message else "000000")
        for col in range(1, len(ANCHOR_COLS) + 1):
            wa.cell(row=arow, column=col).border = BORDER

        for label, hit in ([("LATEST", latest)] if latest else []) + \
                          [(l, anchors.get(l)) for l, _ in ANCHORS]:
            if not hit:
                continue
            for col, val in ((1, code), (2, name), (3, label)):
                wn.cell(row=nav_row, column=col, value=val).font = BASE
            c = wn.cell(row=nav_row, column=4, value=hit[0]); c.number_format = DATE_FMT; c.font = BASE
            c = wn.cell(row=nav_row, column=5, value=hit[1]); c.number_format = NAV_FMT; c.font = BASE
            wn.cell(row=nav_row, column=6, value=source).font = BASE
            for col in range(1, 7):
                wn.cell(row=nav_row, column=col).border = BORDER
            nav_row += 1

        for col, val in enumerate([run_ts, code, name, status, source, records,
                                   latest[0] if latest else None, message], start=1):
            c = wl.cell(row=log_row, column=col, value=val)
            c.font, c.border = BASE, BORDER
            if col == 7 and latest:
                c.number_format = DATE_FMT
            if col == 4:
                c.font = Font(name=FONT, size=10, bold=True,
                              color={"success": "008000", "partial": "BF8F00",
                                     "fail": "C00000"}.get(status, "000000"))
        log_row += 1
        if on_progress:
            on_progress(i + 1, len(codes), code, status, name)
        else:
            print(f"  [{status:7}] {code:>8}  {str(name)[:56]}")

    write_returns_formulas(wr, len(codes))
    return wb, ok, fail


def refresh(path: str, allow_amfi_history: bool = False):
    """CLI entry point: load from disk, refresh, save back to disk."""
    wb = openpyxl.load_workbook(path)
    wb, ok, fail = refresh_workbook(wb, allow_amfi_history=allow_amfi_history)
    wb.save(path)
    print(f"\n{ok} ok / {fail} failed -> {path}")
    return ok, fail


# ---- offline self-test -----------------------------------------------------

def selftest():
    fails = []

    def check(label, got, want, tol=1e-9):
        good = abs(got - want) <= tol if isinstance(want, float) else got == want
        print(f"  {'PASS' if good else 'FAIL'}  {label}: got {got!r}, want {want!r}")
        if not good:
            fails.append(label)

    start, series, nav = date(2020, 1, 1), [], 100.0
    d = start
    while d <= date(2026, 8, 21):
        if d.weekday() < 5:
            series.append((d, round(nav, 5)))
            nav *= 1.0003
        d += timedelta(days=1)

    print("\n-- month arithmetic --")
    check("31 Mar minus 1M clamps to Feb", months_back(date(2026, 3, 31), 1), date(2026, 2, 28))
    check("21 Aug 2026 minus 36M", months_back(date(2026, 8, 21), 36), date(2023, 8, 21))

    print("\n-- anchor carry-back --")
    check("Sunday target -> Friday", pick_anchor(series, date(2026, 8, 23))[0], date(2026, 8, 21))
    check("never carries forward", pick_anchor(series, date(2020, 1, 4))[0], date(2020, 1, 3))
    check("younger than period -> None", pick_anchor(series, date(2019, 1, 1)), None)

    print("\n-- returns vs independent calc --")
    anc = build_anchor_set(series)
    latest_d, latest_n = anc["latest"]
    d1m, n1m = anc["1M"]
    weekdays = sum(1 for dd, _ in series if d1m < dd <= latest_d)
    check("1M return matches weekday-count expectation",
          abs((latest_n / n1m - 1) - (1.0003 ** weekdays - 1)) < 1e-6, True)
    d3y, n3y = anc["3Y"]
    days = (latest_d - d3y).days
    cagr = (latest_n / n3y) ** (365.25 / days) - 1
    approx = (1 + (1.0003 ** (days * 5 / 7) - 1)) ** (365.25 / days) - 1
    print(f"  3Y {d3y}->{latest_d} ({days}d): formula {cagr:.4%} vs approx {approx:.4%}")
    check("CAGR within 0.5pp of approximation", abs(cagr - approx) < 0.005, True)

    print("\n-- AMFI dual-format parser --")
    old = "119551;INF209KA12Z1;INF209KA13Z9;ABSL Banking & PSU Debt Fund;106.8821;21-Aug-2026"
    new = ("119551;INF209KA12Z1;INF209KA13Z9;Aditya Birla Sun Life Banking & PSU Debt Fund;"
           "Direct Plan;IDCW-Re-investment;106.8821;21-Aug-2026")
    ro, rn = parse_amfi_line(old), parse_amfi_line(new)
    check("OLD 6-col NAV", ro["nav"], 106.8821)
    check("OLD 6-col date", ro["date"], date(2026, 8, 21))
    check("NEW 8-col NAV", rn["nav"], 106.8821)
    check("NEW 8-col date", rn["date"], date(2026, 8, 21))
    check("NEW 8-col plan", rn["plan"], "Direct Plan")
    check("NEW 8-col option", rn["option"], "IDCW-Re-investment")
    check("zero-NAV row rejected",
          parse_amfi_line("148304;INF090I01VZ8;-;Franklin;Direct Plan;Growth;0.0000;21-Aug-2026"),
          None)
    check("section header rejected",
          parse_amfi_line("Open Ended Schemes(Debt Scheme - Banking and PSU Fund)"), None)
    check("column header rejected",
          parse_amfi_line("Scheme Code;ISIN Div Payout/ ISIN Growth;ISIN Div Reinvestment;"
                          "Scheme Name;Plan;Option;Net Asset Value;Date"), None)

    print("\n-- AMFI 90-day chunking --")
    wins = chunk_ranges(date(2023, 8, 21), date(2026, 8, 21))
    check("3Y needs 13 windows", len(wins), 13)
    check("no window exceeds 90 days",
          max((b - a).days + 1 for a, b in wins) <= AMFI_HISTORY_MAX_DAYS, True)
    check("windows are contiguous",
          all(wins[i][1] + timedelta(days=1) == wins[i + 1][0] for i in range(len(wins) - 1)),
          True)
    check("last window ends on end date", wins[-1][1], date(2026, 8, 21))

    print("\n-- outlier + IDCW detection --")
    dirty = series[:-1] + [(series[-1][0], series[-1][1] * 1.4)]
    check("40% spike flagged", spike_flag(dirty, dirty[-1]), True)
    check("clean point not flagged", spike_flag(series, series[-40]), False)
    check("IDCW detected", is_idcw("ABSL Banking & PSU - Direct Plan - IDCW-Re-investment"), True)
    check("Growth not flagged", is_idcw("Parag Parikh Flexi Cap - Direct Plan - Growth"), False)

    print("\n-- bad-row filtering --")
    parsed = parse_history({"data": [
        {"date": "21-08-2026", "nav": "106.88210"},
        {"date": "20-08-2026", "nav": "0.00000"},
        {"date": "bad-date", "nav": "1.0"},
        {"date": "19-08-2026", "nav": ""},
        {"date": "18-08-2026", "nav": "107.13600"}]})
    check("keeps only usable rows", len(parsed), 2)
    check("sorted oldest-first", parsed[0][0], date(2026, 8, 18))

    print("\n-- scheme code resolver (offline) --")
    check("strips plan/option words for search",
          search_query_terms("Axis Small Cap Direct-G"), "axis small cap")
    check("expands Reg-G shorthand",
          normalize_name("Bandhan Large and Mid Cap Reg-G"),
          "bandhan large and mid cap regular growth")
    check("expands Smallcap/Midcap without spaces",
          normalize_name("HSBC Midcap Direct-G"), "hsbc mid cap direct growth")
    check("Teck is preserved, NOT rewritten to Tech (it's the fund's real name)",
          normalize_name("Quant Teck Direct-G"), "quant teck direct growth")

    print("\n-- query variants (spaced vs one-word, AMCs are inconsistent) --")
    check("Smallcap fund gets both spellings tried",
          search_query_variants("Invesco India Smallcap Reg-G"),
          ["invesco india small cap", "invesco india smallcap"])
    check("Midcap fund gets both spellings tried",
          search_query_variants("Motilal Oswal Midcap Direct-G"),
          ["motilal oswal mid cap", "motilal oswal midcap"])
    check("fund with no cap-word gets exactly one query, no wasted call",
          search_query_variants("Quant Teck Direct-G"), ["quant teck"])

    # Regression: Invesco/Motilal Oswal file schemes as one word ("Smallcap",
    # "Midcap"); mfapi's search is a literal substring match, so only the
    # collapsed-form query would have found these in production.
    def fake_search(q):
        bank = {
            "invesco india smallcap": [
                {"schemeCode": 145139,
                 "schemeName": "Invesco India Smallcap Fund - Regular Plan - Growth"},
                {"schemeCode": 145140,
                 "schemeName": "Invesco India Smallcap Fund - Direct Plan - Growth"},
            ],
            "invesco india small cap": [],   # what production actually hit: empty
        }
        return bank.get(q, [])

    global fetch_mfapi_search
    orig_fetch = fetch_mfapi_search
    fetch_mfapi_search = fake_search
    try:
        report = resolve_names(["Invesco India Smallcap Reg-G"])
    finally:
        fetch_mfapi_search = orig_fetch
    check("merged search recovers the one-word-only match",
          report[0]["status"], "ok")
    check("Regular-Growth ranks correctly among merged candidates",
          report[0]["candidates"][0]["code"], "145139")

    fake_candidates = [
        {"schemeCode": 118278, "schemeName": "Axis Small Cap Fund - Regular Plan - Growth"},
        {"schemeCode": 125354, "schemeName": "Axis Small Cap Fund - Direct Plan - Growth"},
        {"schemeCode": 118279, "schemeName": "Axis Small Cap Fund - Regular Plan - IDCW"},
        {"schemeCode": 999999, "schemeName": "Totally Unrelated Debt Fund - Direct - Growth"},
    ]
    ranked = rank_candidates("Axis Small Cap Direct-G", fake_candidates, top_n=5)
    check("Direct-Growth variant ranks first among plan siblings",
          ranked[0]["code"], "125354")
    check("unrelated fund scores lowest",
          ranked[-1]["code"], "999999")
    check("all four candidates returned (below top_n cap)", len(ranked), 4)

    ranked_reg = rank_candidates("Bandhan Large and Mid Cap Reg-G", [
        {"schemeCode": 118419, "schemeName": "Bandhan Large & Mid Cap Fund - Direct Plan - Growth"},
        {"schemeCode": 100001, "schemeName": "Bandhan Large & Mid Cap Fund - Regular Plan - Growth"},
    ])
    check("Regular-Growth outranks Direct-Growth when Regular requested",
          ranked_reg[0]["code"], "100001")

    print(f"\n{'ALL PASSED' if not fails else str(len(fails)) + ' FAILED: ' + ', '.join(fails)}")
    return 1 if fails else 0


def main():
    p = argparse.ArgumentParser(description="MF NAV & Returns workbook refresher")
    p.add_argument("--file", default=DEFAULT_FILE)
    p.add_argument("--init", action="store_true")
    p.add_argument("--selftest", action="store_true")
    p.add_argument("--healthcheck", action="store_true")
    p.add_argument("--amfi-history", action="store_true",
                   help="allow slow AMFI history fallback (90-day chunks)")
    p.add_argument("--resolve", action="store_true",
                   help="look up scheme codes for fund names (no workbook writes)")
    p.add_argument("--names-file", help="text file, one fund name per line (with --resolve)")
    p.add_argument("--names", help="semicolon-separated fund names (with --resolve)")
    a = p.parse_args()

    if a.selftest:
        sys.exit(selftest())
    if a.healthcheck:
        sys.exit(healthcheck())
    if a.resolve:
        if a.names_file:
            with open(a.names_file, encoding="utf-8") as fh:
                names = [ln.strip() for ln in fh if ln.strip()]
        elif a.names:
            names = [n.strip() for n in a.names.split(";") if n.strip()]
        else:
            sys.exit("--resolve needs --names-file <path> or --names \"a;b;c\"")
        print_resolve_report(resolve_names(names))
        return
    if a.init:
        print(f"Scaffold written: {build_scaffold(a.file)}")
        print("Paste scheme codes into the 'Scheme Codes' tab, then: python refresh_nav.py")
        return
    print(f"Refreshing {a.file} ...")
    try:
        refresh(a.file, allow_amfi_history=a.amfi_history)
    except ValueError as exc:
        sys.exit(str(exc))


if __name__ == "__main__":
    main()
